from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from urllib.parse import quote
from datetime import date
import os
import openpyxl

options = Options()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.add_argument("--profile-directory=Default")
options.add_argument("--user-data-dir=/var/tmp/chrome_user_data")
path = "C:\\Users\\QAIT-0081\\Documents\\correo\\datos.xlsx"
today = date.today()

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
first = True

os.system("")
os.environ["WDM_LOG_LEVEL"] = "0"



class style():
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'


f = open("message.txt", "r")
message = f.read()
f.close()

print(style.YELLOW + '\nThis is your message-')
print(style.GREEN + message)
print("\n" + style.RESET)
message = quote(message)
delay = 30

driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
print('Once your browser opens up sign in to web whatsapp')
driver.get('https://web.whatsapp.com')
for i in range(2, m_row + 1):
	nombre = ''
	date_x = ''
	celNumber = ''
	nombre_obj = sheet_obj.cell(row=i, column=1)
	cel_obj = sheet_obj.cell(row=i, column=2)
	date_obj = sheet_obj.cell(row=i, column=3)
	print("DATOS DEL ARCHIVO...")
	print(nombre_obj.value)
	print(cel_obj.value)
	print(date_obj.value)

	nombre = nombre_obj.value
	celNumber = cel_obj.value
	date_x = date_obj.value
	print("diferencia entre días")
	remaining_days = (date_x.date() - today).days
	print("Los días son: " + str(remaining_days))

	#Validacion de vencimiento de poliza al mes
	if remaining_days == 30 or remaining_days == 31:
		if celNumber == "":
			continue
		print(style.YELLOW + '{}/{} => Sending message to {}.' + style.RESET)
		try:
			url = 'https://web.whatsapp.com/send?phone=' + celNumber + '&text=' + message
			sent = False
			for i in range(3):
				if not sent:
					driver.get(url)
					try:
						click_btn = WebDriverWait(driver, delay).until(
						    EC.element_to_be_clickable((By.XPATH, "//button[@class='_4sWnG']")))
					except Exception as e:
						print(style.RED +
						      f"\nFailed to send message to: {celNumber}, retry ({i+1}/3)")
						print("Make sure your phone and computer is connected to the internet.")
						print("If there is an alert, please dismiss it." + style.RESET)
					else:
						sleep(1)
						click_btn.click()
						sent = True
						sleep(3)
						print(style.GREEN + 'Message sent to: ' + celNumber + style.RESET)
		except Exception as e:
			print(style.RED + 'Failed to send message to ' +
			      celNumber + str(e) + style.RESET)
		driver.close()
