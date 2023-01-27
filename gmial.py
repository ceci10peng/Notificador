import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from urllib.parse import quote
from datetime import date
import os
import openpyxl

class style():
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'

path = "C:\\Users\\QAIT-0081\\Documents\\correo\\excel.xlsx"
today = date.today()

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

sender_email_address = 'cecilia_guerrerope@hotmail.com'
sender_email_password = 'Nachito1091'
#receiver_email_address = 'cecilia_guerrerope@hotmail.com'

email_subject_line = 'CAMBIO DE CONTRASEÑA'

for i in range(2, m_row + 1):
    nombre = '' 
    date_x = ''
    receiver_email_address = ''
    nombre_obj = sheet_obj.cell(row=i, column=1)
    correo_obj = sheet_obj.cell(row=i, column=2)
    date_obj = sheet_obj.cell(row=i, column=3)
    print("DATOS DEL ARCHIVO...")
    print(nombre_obj.value)
    print(correo_obj.value)
    print(date_obj.value)

    nombre = nombre_obj.value
    receiver_email_address = correo_obj.value
    date_x = date_obj.value
    print("diferencia entre días")
    remaining_days = (date_x.date() - today).days
    print("Los días son: " + str(remaining_days))

    # Validacion de vencimiento de poliza al mes
    if remaining_days == 30 or remaining_days == 31:
        if receiver_email_address == "":
            continue
        print(style.YELLOW + '{}/{} => Sending message to {}.' + style.RESET)
        try:

            msg = MIMEMultipart()
            msg['From'] = sender_email_address
            msg['To'] = receiver_email_address
            msg['Subject'] = email_subject_line

            email_body = 'Hola Cecilia, Hubo una solicitud para cambiar su contraseña! Haga clic en el enlace para restablecerla:  \n'
            msg.attach(MIMEText(email_body, 'plain'))

            email_content = msg.as_string()
            server = smtplib.SMTP('smtp-mail.outlook.com:587')
            server.starttls()
            server.login(sender_email_address, sender_email_password)

            server.sendmail(sender_email_address, receiver_email_address, email_content)
            server.quit()
        finally:
            print('Se ha Ejecutado')
